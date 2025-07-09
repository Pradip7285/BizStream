import os
import time
import pandas as pd
import logging
from io import StringIO
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait, Select
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.chrome.service import Service as ChromeService
from selenium.webdriver.common.keys import Keys
from webdriver_manager.chrome import ChromeDriverManager
from datetime import datetime, date
from openpyxl import load_workbook, Workbook
from openpyxl.utils.dataframe import dataframe_to_rows

# Configure logging
logger = logging.getLogger(__name__)

# Constants
today = date.today().strftime("%Y-%m-%d")
EXCEL_PATH = os.path.join(os.path.dirname(__file__), "Depot.xlsx")
filename = "stocks.xlsx"

def navigate(driver):
    """Navigate to the stock reports section with error handling"""
    try:
        logger.info("Navigating to stock reports section...")
        
        # Click on the main menu button
        menu_button = WebDriverWait(driver, 15).until(
            EC.element_to_be_clickable((By.ID, 'ctl00_ImageButton11'))
        )
        menu_button.click()
        time.sleep(5)
        
        # Click on the stock reports link
        stock_link = WebDriverWait(driver, 15).until(
            EC.element_to_be_clickable((By.ID, 'ctl00_ContentPlaceHolder1_TabContainer1_Tab_BI_Module_grid_crop_suppl_ctl09_link_crop_supplier'))
        )
        stock_link.click()
        time.sleep(10)
        
        logger.info("Successfully navigated to stock reports section")
        
    except Exception as e:
        logger.error(f"Navigation failed: {e}")
        raise Exception(f"Failed to navigate to stock reports section: {str(e)}")

def submit_request(driver, destination, download_dir, max_retries=3):
    """Submit stock request with retry logic and comprehensive error handling"""
    filepath = os.path.join(download_dir, filename)
    logger.info(f"Processing depot: {destination} -> {filepath}")

    for attempt in range(max_retries):
        try:
            # Select warehouse
            warehouse_select = WebDriverWait(driver, 15).until(
                EC.presence_of_element_located((By.ID, "ctl00_ContentPlaceHolder1_ddl_warehouse_Name"))
            )
            select = Select(warehouse_select)
            select.select_by_visible_text(destination)
            time.sleep(1)

            # Wait for table to update
            table_elem = WebDriverWait(driver, 15).until(
                EC.visibility_of_element_located((By.ID, "ctl00_ContentPlaceHolder1_Grid_req"))
            )
            old_html = table_elem.get_attribute('outerHTML')
            # Click or trigger the event that loads new data if needed (if not automatic)
            # Wait for the table's HTML to change
            WebDriverWait(driver, 15).until(
                lambda d: d.find_element(By.ID, "ctl00_ContentPlaceHolder1_Grid_req").get_attribute('outerHTML') != old_html
            )
            html = driver.find_element(By.ID, "ctl00_ContentPlaceHolder1_Grid_req").get_attribute('outerHTML')

            # Parse table data
            df = pd.read_html(StringIO(html))[0]
            if df.empty:
                logger.warning(f"No data found for depot: {destination}")
                return False

            # Add depot column
            df.insert(0, "Depot", destination)

            # Save to Excel
            if append_df_to_excel(filepath, df):
                logger.info(f"✅ Data saved successfully for {destination}")
                return True
            else:
                logger.error(f"❌ Failed to save data for {destination}")
                return False

        except Exception as e:
            logger.warning(f"Request attempt {attempt + 1} failed for {destination}: {e}")
            if attempt < max_retries - 1:
                time.sleep(3)  # Wait before retry
            else:
                logger.error(f"Failed to process {destination} after {max_retries} attempts")
                return False

    return False

def append_df_to_excel(filepath, df, sheet_name='Sheet1'):
    """Append dataframe to Excel file with error handling"""
    try:
        # Load existing workbook or create new one
        if os.path.exists(filepath):
            wb = load_workbook(filepath)
            logger.info(f"Loaded existing Excel file: {filepath}")
        else:
            wb = Workbook()
            logger.info(f"Created new Excel file: {filepath}")
        
        # Get or create worksheet
        if sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
        else:
            ws = wb.create_sheet(sheet_name)
            logger.info(f"Created new worksheet: {sheet_name}")
        
        # Determine if we need headers
        include_header = ws.max_row == 1
        
        # Append data
        for r in dataframe_to_rows(df, index=False, header=include_header):
            ws.append(r)
        
        # Save workbook
        wb.save(filepath)
        logger.info(f"Successfully appended {len(df)} rows to {filepath}")
        return True
        
    except Exception as e:
        logger.error(f"Error appending to Excel file {filepath}: {e}")
        return False

def scrape_reports(driver, download_dir):
    """Main stock scraping function with comprehensive error handling"""
    try:
        logger.info("Starting stock reports scraping...")
        
        # Load Excel data
        if not os.path.exists(EXCEL_PATH):
            raise FileNotFoundError(f"Excel file not found: {EXCEL_PATH}")
            
        df = pd.read_excel(EXCEL_PATH)
        logger.info(f"Loaded {len(df)} depot entries from Excel")
        
        # Navigate to stock reports section
        navigate(driver)
        
        # Wait for page to be ready
        WebDriverWait(driver, 15).until(
            lambda d: d.execute_script('return document.readyState') == 'complete'
        )
        
        success_count = 0
        failure_count = 0
        
        # Process each depot
        for index, row in df.iterrows():
            try:
                destination = row["Depot"]
                logger.info(f"Processing depot {index + 1}/{len(df)}: {destination}")
                
                if submit_request(driver, destination, download_dir):
                    success_count += 1
                else:
                    failure_count += 1
                    
            except Exception as e:
                failure_count += 1
                logger.error(f"❌ Error processing depot {destination}: {e}")
                continue
        
        logger.info(f"✅ Stock reports scraping completed. Success: {success_count}, Failures: {failure_count}")
        
        # Don't quit driver here as it's managed by the main bot
        logger.info("Stock reports scraping finished successfully")
        
    except Exception as e:
        logger.error(f"❌ Stock reports scraping failed: {e}")
        raise Exception(f"Stock reports scraping failed: {str(e)}")
